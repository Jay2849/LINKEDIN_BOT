import time
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

HR_KEYWORDS = ["hr", "human resource", "talent", "talent acquisition", "dream job", "job"]


def load_keywords_from_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]


def load_message_template(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()


def search_jobs(driver, keyword):
    driver.get("https://www.linkedin.com/jobs/")
    time.sleep(3)
    
    search_input = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Search jobs"]')
    search_input.clear()
    search_input.send_keys(keyword)
    time.sleep(1)
    search_input.submit()
    time.sleep(5)

    apply_filters(driver)


def apply_filters(driver):
    try:
        driver.find_element(By.XPATH, "//button[contains(., 'Date posted')]").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[contains(., 'Past 24 hours')]").click()
        time.sleep(1)

        driver.find_element(By.XPATH, "//button[contains(., 'On-site/Remote')]").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[contains(., 'Remote')]").click()
        time.sleep(1)

        driver.find_element(By.XPATH, "//button[contains(., 'All filters')]").click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//label[contains(., 'Part-time')]").click()
        driver.find_element(By.XPATH, "//label[contains(., 'Temporary')]").click()
        driver.find_element(By.XPATH, "//button[contains(., 'Show results')]").click()
        time.sleep(5)

    except Exception as e:
        print("Error applying filters:", e)


def get_filtered_job_elements(driver):
    return driver.find_elements(By.CSS_SELECTOR, 'ul.jobs-search__results-list li')


def extract_job_details(driver, job_element):
    try:
        job_element.click()
        time.sleep(3)

        title = driver.find_element(By.CSS_SELECTOR, 'h2.t-24').text
        company = driver.find_element(By.CSS_SELECTOR, 'a.topcard__org-name-link').text
        location = driver.find_element(By.CSS_SELECTOR, 'span.topcard__flavor--bullet').text
        description = driver.find_element(By.CSS_SELECTOR, 'div.description').text
        joburl = driver.current_url

        return {
            "NAME": company,
            "TITLE": title,
            "LOCATION": location,
            "JOBURL": joburl,
            "description": description.lower()
        }
    except Exception as e:
        print("Error extracting job details:", e)
        return None


def is_hr_related(description):
    return any(keyword in description for keyword in HR_KEYWORDS)


def personalize_message(template, details):
    message = template
    for key in ["NAME", "TITLE", "LOCATION", "JOBURL"]:
        message = message.replace(f"[{key}]", details[key])
    return message


def send_message(driver, job_element, message):
    try:
        connect_button = driver.find_element(By.XPATH, "//button[contains(., 'Easy Apply') or contains(., 'Apply') or contains(., 'Connect')]")
        if connect_button:
            print("Sending message:", message)
            # Placeholder for actual sending logic
    except Exception:
        print("No messaging option found, skipping.")
